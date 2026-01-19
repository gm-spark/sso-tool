


"""
Created on Tue Jun 25 09:40:47 2024

@author: james
"""


'''
to run this script
SSO Sales Prod needs to be run and placed in the System/Vendor folder for that vendor
Demand Forecast Prod needs to be run for that vendor also 
Supplier Inventory Available Input must be completed (manually)
'''

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def calculate_order_packs(week_need, whse_pack):
    ratio = week_need / whse_pack
    if ratio < 0.5:
        return 0
    elif 0.5 <= ratio < 1:
        return 1
    else:
        return round(ratio)
    
def get_last_values(values_list, x):
    return values_list[-x:]

def get_next_values(values_list, x):
    return values_list[:x]


FC_stores_to_remove = [747, 2360, 2552, 2666, 2680, 3008,\
                       3069, 3149, 3166, 3524, 3547, 3885,\
                           5195, 5344, 5426, 5625, 5638, 5645, 5646,\
                           5697, 5781, 5899, 5941, 5954, 5980, 6178]
    
'''Parameters
'''
#WOS_targets = [101]
WOS_targets = list(range(1,14))
#WOS_targets = [1,4,8,16,100]


#to use the forecast at all. off means only historical
forecast_trigger = 'on'
#to only use the forecast. on means only forecast will be used. 
forecast_only_trigger = 'off'

Supplier = 'DoctorsChoice'
Department = 29
Report_grouping = 'ELSE'  #unused in report

historical_weeks_to_look_back = 4  #this does not include the current week
forecast_end_week_weeks_out = 4  #this includes the current week

#get current WMT week
os.chdir(r"C:\Users\james\Desktop\Data\System")
cal = pd.read_excel("WMT weeks.xlsx")

#determine date parameters
today = pd.Timestamp.today().normalize()
row = cal[cal['Date'] == today]
curr_WMT_week = row['WMT_week'].values[0]
#Year = row['Year'].values[0]
Week = row['Week'].values[0]
Half = row['Half'].values[0]

week_list = cal['WMT_week'].copy()
week_list = week_list[week_list < curr_WMT_week]
week_list = week_list.drop_duplicates()
week_list = week_list.sort_values(ascending=True)
week_list = week_list.tolist()

LW_list = get_last_values(week_list, 1)
L4W_list = get_last_values(week_list, 4)
L52W_list = get_last_values(week_list, 52)
L10W_list = get_last_values(week_list, 10)
L2W_list = get_last_values(week_list, 2)
lookback_week_list = get_last_values(week_list, historical_weeks_to_look_back)


LW = LW_list[0]
LLW = L2W_list[0]
Year = int(str(LW)[:4])

#do upcoming weeks functionality
week_list = cal['WMT_week'].copy()
week_list = week_list[week_list >= curr_WMT_week]
week_list = week_list.drop_duplicates()
week_list = week_list.sort_values(ascending=True)
week_list = week_list.tolist()

forecast_week_list = get_next_values(week_list, forecast_end_week_weeks_out)


# LW = curr_WMT_week - 1
# forecast_end_week = curr_WMT_week + forecast_end_week_weeks_out

#bring in store sales datasets

os.chdir(fr"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\Reporting Automation\System\Vendors\{Supplier}")
si = pd.read_excel(f"{Supplier} Store Items Prod.xlsx")
#ss = pd.read_csv(f"{Supplier} Store Sales Prod.csv")
ss = pd.read_csv(f"{Supplier} Store Sales Reduced L10W.csv")
#ss = pd.read_csv(f"{Supplier} Store Sales Prod.csv")
ss = ss[ss['omni_department_number']==Department]



#calculate average store item sales

avg_sic_sales = ss[['all_links_item_number', 'store_number','pos_sales_this_year', \
          'pos_quantity_this_year', 'walmart_calendar_week',]].copy()

avg_sic_sales = avg_sic_sales[avg_sic_sales['walmart_calendar_week'].isin(lookback_week_list)]
avg_sic_sales = avg_sic_sales.drop(columns='walmart_calendar_week')
avg_sic_sales = avg_sic_sales.groupby(['all_links_item_number', 'store_number']).mean().reset_index()


#bring in the Supplier Available Dataset
os.chdir(r"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\SSO Tool\Manual Inputs")

Sup_OHs= pd.read_excel(f'{Supplier} Available at Supplier.xlsx')
#Sup_OHs= pd.read_excel(f'{Supplier} Available at Supplier 2.xlsx')
Sup_OHs = Sup_OHs[Sup_OHs['Department Nbr']==Department]
Sup_OHs = Sup_OHs.drop_duplicates(subset='WMT Prime Item Nbr', keep='last')

Sup_OHs['Supplier Has On Hand to Ship'] = Sup_OHs['Supplier Has On Hand to Ship'].apply(pd.to_numeric, errors='coerce')
Sup_OHs = Sup_OHs.dropna(subset='Supplier Has On Hand to Ship')

Sup_OHs['OH Available'] = Sup_OHs['Max quantity we are willing to send right now'].combine_first(Sup_OHs['Supplier Has On Hand to Ship'])
Sup_OHs = Sup_OHs[Sup_OHs['OH Available']>0]
Sup_OHs = Sup_OHs.set_index('WMT Prime Item Nbr')

#bring in SSO dataset
os.chdir(fr"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\Reporting Automation\System\Vendors\{Supplier}")
sso = pd.read_csv(f'{Supplier} SSO Sales Prod.csv') 
sso = sso[~sso['store_number'].isin(FC_stores_to_remove)]
#sso = sso[sso['omni_department_number'==Department]]
sso = sso[sso['traited_store_count_this_year']>0]
#ascertain no duplicates on index we will be joining on 
sso = sso.sort_values(by =['store_on_hand_quantity_this_year'], ascending=False)
sso = sso.drop_duplicates(subset=['all_links_item_number', 'store_number'], keep='first')

sso = pd.merge(sso,avg_sic_sales,how='left',left_on=['all_links_item_number', 'store_number'], \
               right_on=['all_links_item_number', 'store_number'], suffixes=('','_from_sales'))

 
sso = sso.rename(columns={'pos_sales_this_year': 'avg_sales_L8W', \
                         'pos_quantity_this_year' : 'avg_pos_quantity_L8W'})

sso['avg_sales_L8W'] = sso['avg_sales_L8W'].fillna(0.0)
sso['avg_pos_quantity_L8W'] = sso['avg_pos_quantity_L8W'].fillna(0.0)


if forecast_only_trigger == 'on':
    sso['avg_pos_quantity_L8W'] = 0.0
else:
    pass

if forecast_trigger == 'on':
    #bring in demand forecast dataset
    os.chdir(fr"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\Reporting Automation\System\Vendors\{Supplier}")  
    df1 = pd.read_csv(f"{Supplier} Demand Forecast Prod.csv")
    df1 = df1[df1['omni_dept_nbr']==Department]
    #filter to relevant forecast weeks
    df1 = df1[df1['walmart_calendar_week'].isin(forecast_week_list)]
    
    
    df1 = df1[['all_links_item_nbr','store_nbr','final_forecast_each_quantity']]
    df1 = df1.groupby(['all_links_item_nbr','store_nbr']).sum()
    df1['avg_weekly_forecast'] = df1['final_forecast_each_quantity']/forecast_end_week_weeks_out
    df1 = df1.drop(columns = 'final_forecast_each_quantity')
    df1 = df1.reset_index()   
    
    sso = pd.merge(sso, df1, how='left',left_on=['all_links_item_number', 'store_number'], \
                   right_on=['all_links_item_nbr', 'store_nbr'], suffixes=('','_from_sales'))
        
    sso = sso.drop(columns = ['all_links_item_nbr', 'store_nbr'])
    sso['avg_weekly_forecast'] = sso['avg_weekly_forecast'].fillna(0.0)
    sso['avg_pos_quantity_L8W'] = sso[['avg_pos_quantity_L8W', 'avg_weekly_forecast']].max(axis=1)
    
else:
    pass
      
                                         
    
sso['F4 Pipe'] = sso['store_on_hand_quantity_this_year'] + sso['store_in_transit_quantity_this_year'] \
+ sso['store_in_warehouse_quantity_this_year'] + sso['store_on_order_quantity_this_year']


#sso['F4 Pipe'] = 0

df_dict = {}

for x in WOS_targets:
    print(x)
    df = sso.copy()
    df['WOS_target'] = df['avg_pos_quantity_L8W']*x
    
    #for custom edit to not end any store with more than 8 units total pipe
    df['WOS_target'] = df['WOS_target'].apply(lambda x: min(4, x))
    #df['WOS_target'] = 2
    
    df['need'] = df['WOS_target']- df['F4 Pipe']
    
    #for custom edit to set all stores to the lesser of need or a certain pipe level... 
    # df['F4 Pipe'] = df['F4 Pipe'].apply(lambda x: max(0, x))
    # df['need'] = 8 - df['F4 Pipe']    
    # df['need'] = df['need'].apply(lambda x: max(0, x))
    # df['need2'] = df['WOS_target'] - df['F4 Pipe']
    # df['need'] = df[['need', 'need2']].min(axis=1)
    # df = df.drop(columns='need2')        
        
    week_need_col = 'need'
    df['order_packs'] = df.apply(lambda row: calculate_order_packs(row[week_need_col], row['warehouse_pack_quantity']), axis=1)
        
    df['order_eaches'] = df['order_packs']*df['warehouse_pack_quantity']
    
    df['Whse Pack Override'] = df['F4 Pipe'].apply(lambda x: 1 if x <= 1 else 0)
    
    df['Whse Packs Distro'] = df[['order_packs', 'Whse Pack Override']].max(axis=1)
    df['Eaches Distro'] = df['Whse Packs Distro']*df['warehouse_pack_quantity']
    df['week_variable'] = x
    #df = df[['all_links_item_number', 'store_number', 'Whse Packs Distro','Eaches Distro','week_variable']]
    df_dict[f'df_{x}'] = df


combo = pd.concat(df_dict.values(), ignore_index=True)

grouped = combo.groupby(['all_links_item_number', 'week_variable']).sum().reset_index()
grouped = grouped.drop(columns='store_number')
grouped = grouped.join(Sup_OHs, how='left', on='all_links_item_number')
grouped['surplus'] = grouped['Max quantity we are willing to send right now'] - grouped['Eaches Distro']
grouped = grouped[['all_links_item_number', 'week_variable', 'surplus']]

combo = pd.merge(combo, grouped, how='left', left_on=['all_links_item_number', 'week_variable'], \
                 right_on=['all_links_item_number', 'week_variable'])
    
retails = sso[['all_links_item_number', 'base_unit_retail_amount']].copy()
retails = retails.drop_duplicates(subset='all_links_item_number', keep='first')
retails = retails.set_index('all_links_item_number')
combo = combo.drop(columns = 'base_unit_retail_amount')
combo = combo.join(retails, on='all_links_item_number', how='left')
combo['base_unit_retail_amount'] = combo['base_unit_retail_amount'].replace({'\$': ''}, regex=True)
combo['base_unit_retail_amount'] = combo['base_unit_retail_amount'].replace({'\,': ''}, regex=True)
combo['base_unit_retail_amount'] = combo['base_unit_retail_amount'].apply(pd.to_numeric, errors='raise')

rollup_dict = {}

for x in WOS_targets:
    print(x)
    splitout = combo.copy()
    splitout = splitout[splitout['week_variable']<=x]
    splitout = splitout[splitout['surplus']>0]
    splitout = splitout.sort_values('week_variable', ascending=False)
    splitout = splitout.drop_duplicates(subset = ['all_links_item_number', 'store_number'], keep='first')
    splitout = splitout[splitout['Whse Packs Distro']>0]
    splitout['Retail_amount_sent'] = splitout['Eaches Distro']* splitout['base_unit_retail_amount']
    splitout = splitout.sort_values(by=['all_links_item_number', 'store_number'], ascending=True)
    splitout['blank'] = ''
    NOVA = splitout[['all_links_item_number', 'store_number','blank','Whse Packs Distro',\
                  'Eaches Distro','Retail_amount_sent', 'week_variable']].copy()
        
        
    rollup = NOVA[['all_links_item_number', 'store_number', 'Whse Packs Distro',\
                  'Eaches Distro','Retail_amount_sent', 'week_variable']].copy()
        
    agg_dict = {
        'store_number': 'count', 'Whse Packs Distro': 'sum', 'Eaches Distro': 'sum',\
        'Retail_amount_sent': 'sum', 'week_variable': 'mean'}
    
    rename_dict = {
        'store_number': 'store count', 'Whse Packs Distro': 'Total Whse Packs',\
        'Eaches Distro': 'Total Eaches', 'Retail_amount_sent': 'Total Retail Sent',\
        'week_variable': 'Max WOS sent'}
    
    rollup = rollup.groupby('all_links_item_number').agg(agg_dict).reset_index()
    rollup = rollup.rename(columns=rename_dict)
    rollup['from which WOS distro'] = x
    rollup_dict[f'df_{x}'] = rollup
    
    os.chdir(r"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\SSO Tool\Templates")
    wb = load_workbook('NOVA template.xlsx')
    
    ws = wb['NOVA form']
    last_row_idx = 1
    for r_idx, row in enumerate(dataframe_to_rows(NOVA, index=False, header=False), start=2): # Assuming row 2 is first empty row below headers
          last_row_idx = r_idx    
          for c_idx, value in enumerate(row, start=1):  # Assuming column 1 (A) is the start
            ws.cell(row=r_idx, column=c_idx, value=value)
            
    os.chdir(fr"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\SSO Tool\Output\{Supplier}\NOVA format")
    output_path = f"{curr_WMT_week}_{Supplier}_Dept{Department}_{x}WOS.xlsx"
    wb.save(output_path)
    

total_rollup = pd.concat(rollup_dict.values(), ignore_index=True)
#si = si[['all_links_item_number', 'item_name']]
#si = si.set_index('all_links_item_number')
#total_rollup= total_rollup.join(si, how='left', on='all_links_item_number')
total_rollup = total_rollup.melt(id_vars = ['all_links_item_number', 'from which WOS distro'])

os.chdir(fr"C:\Users\james\Spark Strategy Dropbox\Spark Strategy Team Folder\SSO Tool\Output\{Supplier}\Rollup format")

total_rollup.to_excel(f'{curr_WMT_week}_{Supplier}_Dept{Department}_all_runs_summary.xlsx')


                                 
